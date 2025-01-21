import os
import subprocess, sys
import openpyxl
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QBrush, QIcon, QColor
from PyQt5.QtWidgets import QAbstractItemView, QTableWidgetItem, QHeaderView, QPushButton, QTableWidgetSelectionRange, \
    QMessageBox, QLabel
from PyQt5 import QtWidgets, QtCore
from sqlalchemy import or_
from sqlalchemy.sql.functions import user, coalesce
from functools import partial
from Model.database import session
from Model.model import User, Department, Branch, User_System, SziFileInst, SziAccounting
from View.main_container.container import Ui_MainWindow
from ViewModel.Helper_all import Helper_all
from ViewModel.User.user_new import User_new
from ViewModel.main_load import Main_load
from ViewModel.setting_view import Setting_view
from config_helper import config
from stylesheet import style


class User_main(QtWidgets.QMainWindow):
    def __init__(self, path_helper):
        super(User_main, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.s = session()

        self.path_helper = path_helper
        self.setStyleSheet(style)

        width = QtWidgets.qApp.desktop().availableGeometry(self).width()  # Устанавливаем размер долей рабочих панелей
        self.ui.splitter.setSizes([width * 1 / 3, width * 2 / 3])

        self.setWindowTitle("Пользователи")
        self.setWindowIcon(QIcon(self.path_helper + '/Icons/user.png'))

        self.create_btn_info()
        self.ui.verticalLayout_2.addWidget(self.btn_info)

        self.create_tW_info()
        self.ui.verticalLayout_2.addWidget(self.tW_info)

        self.ui.lineEdit_searchUser.textChanged.connect(partial(self.searchUser))

        Main_load.create_tW_list(self.ui)
        self.ui.tW_list.itemSelectionChanged.connect(self.changed_current_cell_user)
        self.ui.tW_list.itemDoubleClicked.connect(self.clicked_btn_edit)
        Main_load.print_list(self.ui, self.load_user())

        Main_load.select_row_intable(self.ui)

        self.BUTTON_DISPLAY = ('ФИО:', 'И.О. Ф:', 'Ф И.О.:')
        self.fio = ''

        self.create_menu_button()

        self.verticalSpacer = QtWidgets.QSpacerItem(20, 40, QtWidgets.QSizePolicy.Minimum,
                                                    QtWidgets.QSizePolicy.Expanding)  # Подпружиниваем по вертикали
        self.ui.verticalLayout_2.addItem(self.verticalSpacer)

    def searchUser(self):

        Main_load.print_list(self.ui, self.load_user())
        Main_load.select_row_intable(self.ui)

    def clicked_fio(self):
        '''При каждом нажатии на ФИО меняется отображение ФИО'''
        row = self.tW_info.currentRow()
        column = self.tW_info.currentColumn()

        if row == 3 and column == 0:
            text = self.tW_info.item(row, column).text()
            index = self.BUTTON_DISPLAY.index(text)
            if index >= 2:
                index = 0
            else:
                index += 1

            self.tW_info.setItem(3, 0, QTableWidgetItem(self.BUTTON_DISPLAY[index]))

            self.tW_info.item(row, column).setBackground(QBrush(Qt.lightGray))  # серый фон
            self.tW_info.item(row, column).setTextAlignment(
                Qt.AlignVCenter | Qt.AlignRight)  # Выравнивание по правому краю
            self.tW_info.item(row, column).setFlags(
                Qt.ItemIsEnabled | Qt.ItemIsSelectable)  # запрет редактирования

            collor_tex = QColor('#07437D')
            self.tW_info.item(3, 0).setForeground(QBrush(collor_tex))

            self.display_fio()

    def create_tW_info(self):
        '''Создаем табицу инфо tW_info'''
        self.tW_info = QtWidgets.QTableWidget()
        self.tW_info.setShowGrid(False)
        self.tW_info.setWordWrap(True)
        self.tW_info.setCornerButtonEnabled(True)


        numrows = 12
        numcols = 2

        self.tW_info.setColumnCount(numcols)
        self.tW_info.setRowCount(numrows)

        header = self.tW_info.horizontalHeader()
        header.setSectionResizeMode(1, QHeaderView.Stretch)
        self.tW_info.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)

        self.tW_info.horizontalHeader().setVisible(False)
        self.tW_info.verticalHeader().setVisible(False)

        self.tW_info.clicked.connect(self.clicked_fio)

        self.tW_info.setItem(0, 0, QTableWidgetItem('Филиал:'))
        self.tW_info.setItem(1, 0, QTableWidgetItem('Служба:'))
        self.tW_info.setItem(2, 0, QTableWidgetItem('Табельный №:'))
        self.tW_info.setItem(3, 0, QTableWidgetItem('ФИО:'))
        self.tW_info.setItem(4, 0, QTableWidgetItem('Должность:'))
        self.tW_info.setItem(5, 0, QTableWidgetItem('Телефон:'))
        self.tW_info.setItem(6, 0, QTableWidgetItem('Почта:'))
        self.tW_info.setItem(7, 0, QTableWidgetItem('Расположение:'))
        self.tW_info.setItem(8, 0, QTableWidgetItem('Логин:'))
        self.tW_info.setItem(9, 0, QTableWidgetItem('ДК:'))
        self.tW_info.setItem(10, 0, QTableWidgetItem('АРМ:'))
        self.tW_info.setItem(11, 0, QTableWidgetItem('Работает:'))

        height_tW = Main_load.get_height_qtable(numrows)
        self.tW_info.setMaximumHeight(height_tW - 22)
        self.tW_info.setMinimumHeight(height_tW - 22)

        self.checkBox_status = QtWidgets.QCheckBox()
        self.tW_info.setCellWidget(11, 1, self.checkBox_status)
        self.checkBox_status.setEnabled(False)

        for rowPosition in range(numrows):
            self.tW_info.item(rowPosition, 0).setBackground(QBrush(Qt.lightGray))  # серый фон
            self.tW_info.item(rowPosition, 0).setTextAlignment(
                Qt.AlignVCenter | Qt.AlignRight)  # Выравнивание по правому краю
            self.tW_info.item(rowPosition, 0).setFlags(
                Qt.ItemIsEnabled | Qt.ItemIsSelectable)  # запрет редактирования

        table = QTableWidgetItem()
        table.setTextAlignment(3)

        self.tW_info.setEditTriggers(QAbstractItemView.AllEditTriggers)
        self.tW_info.setFocusPolicy(Qt.NoFocus)
        self.tW_info.setSelectionMode(QAbstractItemView.NoSelection)

        collor_tex = QColor('#07437D')
        self.tW_info.item(3, 0).setForeground(QBrush(collor_tex))

    def clicked_btn_info(self):
        if Main_load.get_id(self.ui) == None:
            return

        if self.btn_info.isChecked():
            self.btn_info.setIcon(QIcon(self.path_helper + '/Icons/unwrap.png'))
            self.tW_info.setVisible(True)
            self.print_fill_user(Main_load.get_id(self.ui))
        else:
            self.btn_info.setIcon(QIcon(self.path_helper + '/Icons/wrap.png'))
            self.tW_info.setVisible(False)

    def create_btn_info(self):
        '''Создаем кноку свернуть развернуть инфо'''
        self.btn_info = QtWidgets.QPushButton()
        self.btn_info.clicked.connect(self.clicked_btn_info)
        self.btn_info.setText('Информация о пользователе')
        self.btn_info.setCheckable(True)
        self.btn_info.setChecked(True)
        self.btn_info.setIcon(QIcon(self.path_helper + '/Icons/unwrap.png'))
        self.btn_info.setIconSize(QtCore.QSize(30, 30))

    def create_menu_button(self):
        '''Создаем МенюБар и кноки в меню'''
        self.ui.action_exit.triggered.connect(self.close)
        self.ui.action_add.triggered.connect(self.clicked_btn_new)
        self.ui.action_edit.triggered.connect(self.clicked_btn_edit)
        self.ui.action_sort.triggered.connect(self.clicked_btn_sort)
        self.ui.action_export.triggered.connect(self.clicked_btn_export)

        self.ui.btn_new.clicked.connect(self.clicked_btn_new)
        self.ui.btn_edit.clicked.connect(self.clicked_btn_edit)
        self.ui.btn_sort.clicked.connect(self.clicked_btn_sort)
        self.ui.btn_export.clicked.connect(self.clicked_btn_export)

        self.ui.btn_new.setIcon(QIcon(self.path_helper + '/Icons/User/add-user-male-40.png'))
        self.ui.btn_edit.setIcon(QIcon(self.path_helper + '/Icons/User/registration-40.png'))
        self.ui.btn_export.setIcon(QIcon(self.path_helper + '/Icons/microsoft-excel-40.png'))
        self.ui.btn_sort.setIcon(QIcon(self.path_helper + '/Icons/User/search-user-40.png'))

        self.ui.btn_new.setIcon(QIcon(self.path_helper + '/Icons/add.png'))
        self.ui.btn_edit.setIcon(QIcon(self.path_helper + '/Icons/edit.png'))
        self.ui.btn_export.setIcon(QIcon(self.path_helper + '/Icons/export.png'))
        self.ui.btn_sort.setIcon(QIcon(self.path_helper + '/Icons/sorts.png'))

    def clicked_btn_new(self):
        self.new_user = User_new()
        self.new_user.exec_()

        Main_load.print_list(self.ui, self.load_user())
        Main_load.select_row_intable(self.ui, str(self.new_user.employeeId))

    def clicked_btn_edit(self):
        employeeId = Main_load.get_id(self.ui)
        if employeeId is not None:
            self.new_user = User_new(employeeId)
            self.new_user.exec_()

            Main_load.print_list(self.ui, self.load_user())
            Main_load.select_row_intable(self.ui, str(self.new_user.employeeId))

    def clicked_btn_sort(self):
        self.settingViewUser = Setting_view('user')
        self.settingViewUser.exec_()

        Main_load.print_list(self.ui, self.load_user())
        Main_load.select_row_intable(self.ui)

    def clicked_btn_export(self):
        '''Экспорт пользователей в эксель'''
        result = QMessageBox.warning(self, 'Предупреждение',
                                     "Вы действительно хотите выгрузить пользователь в XLSX?",
                                     QMessageBox.Ok, QMessageBox.Cancel)
        if result == 1024:
            count_row = self.ui.tW_list.rowCount()
            list_employeeId = []
            for i in range(count_row):
                list_employeeId.append(self.ui.tW_list.item(i, 0).text())
            list_export_users = []
            for i in list_employeeId:
                list_export_users.append(
                    self.s.query(User.employeeId, User.fio, Department.name, User.post, User.phone, User.eMail,
                                 User.address, User.login, User.dk, User.armName).select_from(User).join(
                        Department).filter(
                        User.employeeId == i).one())

            wb = openpyxl.Workbook()
            sheet = wb.sheetnames
            lis = wb.active
            # Создание строки с заголовками
            lis.append(('Табельный номер', 'ФИО', 'Служба', 'Должность', 'Телефон', 'Почта', 'Расположение', 'Логин',
                        'Договор конфиденциальности', 'Доступные АРМ'))
            for user in list_export_users:
                lis.append(list(user))

                path_export = self.path_helper + '/' + 'Export_files' + '/'
            if not os.path.exists(path_export):
                os.makedirs(path_export)

            wb.save(filename=path_export + 'users.xlsx')
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.call([opener, path_export + '/' + 'users.xlsx'])

    def display_fio(self):
        '''Функция изменения отображения ФИО'''
        if self.fio != None:
            index = self.tW_info.item(3, 0).text()
            self.tW_info.setItem(3, 1, QTableWidgetItem(Helper_all.display_fio(self.fio, index)))

    def print_fill_user(self, employeeId):
        user = self.s.query(User, Department.name, Branch.name). \
            join(Department).join(Branch). \
            filter(User.employeeId == employeeId). \
            one()

        self.tW_info.setItem(0, 1, QTableWidgetItem(user[2]))
        self.tW_info.setItem(1, 1, QTableWidgetItem(user[1]))
        self.tW_info.setItem(2, 1, QTableWidgetItem(str(user[0].employeeId)))

        self.fio = user[0].fio
        index = self.tW_info.item(3, 0).text()

        self.display_fio()
        self.tW_info.setItem(4, 1, QTableWidgetItem(user[0].post))
        self.tW_info.setItem(5, 1, QTableWidgetItem(user[0].phone))
        self.tW_info.setItem(6, 1, QTableWidgetItem(user[0].eMail))
        self.tW_info.setItem(7, 1, QTableWidgetItem(user[0].address))
        self.tW_info.setItem(8, 1, QTableWidgetItem(user[0].login))
        self.tW_info.setItem(9, 1, QTableWidgetItem(user[0].dk))
        self.tW_info.setItem(10, 1, QTableWidgetItem(user[0].armName))
        self.checkBox_status.setChecked(user[0].statusWork)

    def print_tW_info(self, id):
        if id is not None:
            if self.btn_info.isChecked():
                self.print_fill_user(Main_load.get_id(self.ui))

            # if self.btn_notes.isChecked():
            #     self.print_tE_notes(user_id)
            #
            # if self.btn_szi.isCheckable():
            #     self.print_tW_Szi(user_id)

        else:
            self.tW_info.setItem(0, 1, QTableWidgetItem(''))
            self.tW_info.setItem(1, 1, QTableWidgetItem(''))
            self.tW_info.setItem(2, 1, QTableWidgetItem(''))
            self.tW_info.setItem(3, 1, QTableWidgetItem(''))
            self.tW_info.setItem(4, 1, QTableWidgetItem(''))
            self.tW_info.setItem(5, 1, QTableWidgetItem(''))
            self.tW_info.setItem(6, 1, QTableWidgetItem(''))
            self.tW_info.setItem(7, 1, QTableWidgetItem(''))
            self.tW_info.setItem(8, 1, QTableWidgetItem(''))
            self.tW_info.setItem(9, 1, QTableWidgetItem(''))
            self.tW_info.setItem(10, 1, QTableWidgetItem(''))
            self.tW_info.setItem(11, 1, QTableWidgetItem(''))

    def changed_current_cell_user(self):
        '''Действие при смене ячейки в списке'''
        user_id = Main_load.get_id(self.ui)
        self.print_tW_info(user_id)

        # self.tE_notes.setText('')
        #
        # self.ui.tW_Szi.setRowCount(0)

        # self.create_table()
        #
        # self.print_tW_user()
        #
        # self.path_helper = path_helper
        #
        # self.btn_newUser()
        # self.btn_editUser()
        # self.btn_settingViewUser()
        # self.btn_expotyExl()
        #
        # self.ui.btn_info.clicked.connect(self.clicked_btn_info)
        # self.ui.btn_info.setText('Информация пользователя')
        # self.ui.btn_info.setIcon(QIcon(self.path_helper + '/Icons/unwrap.png'))
        # self.ui.btn_info.setIconSize(QtCore.QSize(30, 30))
        # self.btn_info_toggled = False
        #
        #
        # self.ui.tW_list.itemDoubleClicked.connect(self.clicked_btn_editUser)
        #
        # self.BUTTON_DISPLAY = ('ФИО:', 'И.О. Ф:', 'Ф И.О.:')
        # self.fio=''
        #
        # self.ui.lineEdit_searchUser.textChanged.connect(partial(self.print_tW_user, None))
        #
        # horizSpacer = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Minimum)
        # self.ui.horizontalLayout_4.addItem(horizSpacer)
        #
        # width = QtWidgets.qApp.desktop().availableGeometry(self).width()
        # self.ui.splitter.setSizes([width * 1 / 3, width * 2 / 3])

    def load_user(self):
        '''Запрос в SQL список пользователей'''

        checked_branch = config['User']['checked_branch']
        checked_branch = list(checked_branch)
        # print(checked_branch)

        checked_department = config['User']['checked_department']
        checked_department = list(checked_department)
        # print(checked_department)

        checked_szi = list(map(int, config['User']['checked_szi']))
        if Helper_all.convert_bool(config['User']['checkBox_szi']) == True:
            checked_szi.append(0)

        checked_system = list(map(int, config['User']['checked_system']))
        if Helper_all.convert_bool(config['User']['checkBox_system']) == True:
            checked_system.append(0)

        serch_text = '%' + self.ui.lineEdit_searchUser.text() + '%'
        if serch_text == '':
            serch_text = '%'

        status_work = Helper_all.get_status(Helper_all.convert_bool(config['User']['checkB_statusOn']),
                                            Helper_all.convert_bool(config['User']['checkB_statusOff']))

        # else:
        user = self.s.query(User.employeeId, User.fio).select_from(User). \
            join(User_System, isouter=True). \
            join(SziFileInst, isouter=True). \
            join(SziAccounting, isouter=True). \
            filter(or_(
            coalesce(SziAccounting.sziType_id, 0).in_((checked_szi))),
            coalesce(SziAccounting.status, True) == True). \
            filter(coalesce(User_System.id_inf_system, 0).in_((checked_system))). \
            filter(User.branch_id.in_((checked_branch))). \
            filter(User.departmet_id.in_((checked_department))). \
            filter(User.statusWork.like(status_work)). \
            filter(or_(
            User.employeeId.like(serch_text),
            User.fio.like(serch_text),
            User.post.like(serch_text),
            User.phone.like(serch_text),
            User.eMail.like(serch_text),
            User.address.like(serch_text),
            User.login.like(serch_text),
            User.dk.like(serch_text),
            User.armName.like(serch_text))). \
            group_by(User.employeeId). \
            order_by(User.fio)

        return user
        # user = self.s.query(User.employeeId, User.fio).select_from(User). \
        #     join(User_System, isouter=True). \
        #     join(SziFileInst, isouter=True). \
        #     join(SziAccounting, isouter=True). \
        #     group_by(User.employeeId). \
        #     order_by(User.fio)
        # return user

    # def clicked_btn_info(self):
    #     if self.btn_info_toggled:
    #         self.ui.btn_info.setIcon(QIcon(self.path_helper + '/Icons/unwrap.png'))
    #         self.btn_info_toggled = False
    #         self.ui.tW_info.setVisible(True)
    #     else:
    #         self.ui.btn_info.setIcon(QIcon(self.path_helper + '/Icons/wrap.png'))
    #         self.btn_info_toggled = True
    #         self.ui.tW_info.setVisible(False)
    #
    # def btn_newUser(self):
    #     btn_newUser = QPushButton('     Добавить\n пользователя')
    #     self.ui.horizontalLayout_4.addWidget(btn_newUser)
    #     btn_newUser.setIcon(QIcon(self.path_helper + '/Icons/User/add-user-male-40.png'))
    #     btn_newUser.setIconSize(QtCore.QSize(40, 40))
    #     btn_newUser.clicked.connect(self.clicked_btn_newUser)
    #
    # def btn_editUser(self):
    #     btn_editUser = QPushButton('Редактировать\n пользователя')
    #     self.ui.horizontalLayout_4.addWidget(btn_editUser)
    #     btn_editUser.setIcon(QIcon(self.path_helper + '/Icons/User/registration-40.png'))
    #     btn_editUser.setIconSize(QtCore.QSize(40, 40))
    #     btn_editUser.clicked.connect(self.clicked_btn_editUser)
    #
    # def btn_settingViewUser(self):
    #     btn_settingViewUser = QPushButton('Сортировка\n  и фильтр')
    #     self.ui.horizontalLayout_4.addWidget(btn_settingViewUser)
    #     btn_settingViewUser.setIcon(QIcon(self.path_helper + '/Icons/User/search-user-40.png'))
    #     btn_settingViewUser.setIconSize(QtCore.QSize(40, 40))
    #     btn_settingViewUser.clicked.connect(self.clicked_settingViewUser)
    #
    # def btn_expotyExl(self):
    #     btn_expotyExl = QPushButton(' Экспорт \n    XLSX')
    #     self.ui.horizontalLayout_4.addWidget(btn_expotyExl)
    #     btn_expotyExl.setIcon(QIcon(self.path_helper + '/Icons/microsoft-excel-40.png'))
    #     btn_expotyExl.setIconSize(QtCore.QSize(40, 40))
    #     btn_expotyExl.clicked.connect(self.clicked_excel)
    #
    # def clicked_excel(self):
    #     '''Экспорт пользователей в эксель'''
    #     result = QMessageBox.warning(self, 'Предупреждение',
    #                                  "Вы действительно хотите выгрузить пользователь в XLSX?",
    #                                  QMessageBox.Ok, QMessageBox.Cancel)
    #     if result == 1024:
    #         count_row = self.ui.tW_list.rowCount()
    #         list_employeeId = []
    #         for i in range(count_row):
    #             list_employeeId.append(self.ui.tW_list.item(i, 0).text())
    #         list_export_users = []
    #         for i in list_employeeId:
    #             list_export_users.append(
    #                 self.s.query(User.employeeId, User.fio, Department.name, User.post, User.phone, User.eMail,
    #                              User.address, User.login, User.dk, User.armName).select_from(User).join(
    #                     Department).filter(
    #                     User.employeeId == i).one())
    #
    #         wb = openpyxl.Workbook()
    #         sheet = wb.sheetnames
    #         lis = wb.active
    #         # Создание строки с заголовками
    #         lis.append(('Табельный номер', 'ФИО', 'Служба', 'Должность', 'Телефон', 'Почта', 'Расположение', 'Логин',
    #                     'Договор конфиденциальности', 'Доступные АРМ'))
    #         for user in list_export_users:
    #             lis.append(list(user))
    #
    #             path_export = self.path_helper + '/' + 'Export_files' + '/'
    #         if not os.path.exists(path_export):
    #             os.makedirs(path_export)
    #
    #         wb.save(filename=path_export + 'users.xlsx')
    #         opener = "open" if sys.platform == "darwin" else "xdg-open"
    #         subprocess.call([opener, path_export + '/' + 'users.xlsx'])
    #
    # def clicked_settingViewUser(self):
    #     self.settingViewUser = Setting_view('user')
    #     self.settingViewUser.exec_()
    #     self.print_tW_user()
    #
    # def clicked_btn_editUser(self):
    #     employeeId = self.get_employeeId()
    #     if employeeId is not None:
    #         self.new_user = User_new(employeeId)
    #         self.new_user.exec_()
    #         self.print_tW_user(int(employeeId))
    #
    # def clicked_btn_newUser(self):
    #     self.new_user = User_new()
    #     self.new_user.exec_()
    #     employeeId = self.get_employeeId()
    #     try:
    #         employeeId = self.new_user.employeeId
    #     except:
    #         employeeId = employeeId
    #     self.print_tW_user(employeeId)
    #
    # def get_employeeId(self):
    #     if self.ui.tW_list.item(self.ui.tW_list.currentRow(), 0) is None:
    #         return None
    #     else:
    #         return self.ui.tW_list.item(self.ui.tW_list.currentRow(), 0).text()
    #
    # def print_fill_user(self, employeeId):
    #     user = self.s.query(User, Department.name, Branch.name). \
    #         join(Department).join(Branch). \
    #         filter(User.employeeId == employeeId). \
    #         one()
    #
    #     self.ui.tW_info.setItem(0, 1, QTableWidgetItem(user[2]))
    #     self.ui.tW_info.setItem(1, 1, QTableWidgetItem(user[1]))
    #     self.ui.tW_info.setItem(2, 1, QTableWidgetItem(str(user[0].employeeId)))
    #
    #     self.fio=user[0].fio
    #     index = self.ui.tW_info.item(3, 0).text()
    #
    #     self.display_fio()
    #     self.ui.tW_info.setItem(4, 1, QTableWidgetItem(user[0].post))
    #     self.ui.tW_info.setItem(5, 1, QTableWidgetItem(user[0].phone))
    #     self.ui.tW_info.setItem(6, 1, QTableWidgetItem(user[0].eMail))
    #     self.ui.tW_info.setItem(7, 1, QTableWidgetItem(user[0].address))
    #     self.ui.tW_info.setItem(8, 1, QTableWidgetItem(user[0].login))
    #     self.ui.tW_info.setItem(9, 1, QTableWidgetItem(user[0].dk))
    #     self.ui.tW_info.setItem(10, 1, QTableWidgetItem(user[0].armName))
    #     self.checkBox_status.setChecked(user[0].statusWork)
    #
    # def changed_current_cell_user(self):
    #
    #     user_id = self.get_employeeId()
    #     if user_id is not None:
    #         if self.btn_info_toggled == False:
    #             self.print_fill_user(user_id)
    #     else:
    #         self.ui.tW_info.setItem(0, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(1, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(2, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(3, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(4, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(5, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(6, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(7, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(8, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(9, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(10, 1, QTableWidgetItem(''))
    #         self.ui.tW_info.setItem(11, 1, QTableWidgetItem(''))
    #
    # def clicked_fio(self):
    #     '''При каждом нажатии на ФИО меняется отображение ФИО'''
    #     row = self.ui.tW_info.currentRow()
    #     column = self.ui.tW_info.currentColumn()
    #
    #     if row == 3 and column == 0:
    #         text = self.ui.tW_info.item(row, column).text()
    #         index = self.BUTTON_DISPLAY.index(text)
    #         if index >= 2:
    #             index = 0
    #         else:
    #             index += 1
    #
    #         self.ui.tW_info.setItem(3, 0, QTableWidgetItem(self.BUTTON_DISPLAY[index]))
    #
    #         self.ui.tW_info.item(row, column).setBackground(QBrush(Qt.lightGray))  # серый фон
    #         self.ui.tW_info.item(row, column).setTextAlignment(
    #             Qt.AlignVCenter | Qt.AlignRight)  # Выравнивание по правому краю
    #         self.ui.tW_info.item(row, column).setFlags(
    #             Qt.ItemIsEnabled | Qt.ItemIsSelectable)  # запрет редактирования
    #
    #         collor_tex = QColor('#07437D')
    #         self.ui.tW_info.item(3, 0).setForeground(QBrush(collor_tex))
    #
    #         self.display_fio()
    #
    # def display_fio(self):
    #     '''Функция изменения отображения ФИО'''
    #     if self.fio != None:
    #         index = self.ui.tW_info.item(3, 0).text()
    #         self.ui.tW_info.setItem(3, 1, QTableWidgetItem(Helper_all.display_fio(self.fio, index)))
    #
    # def create_table(self):
    #     def create_item_data():
    #         numrows = 12
    #         numcols = 2
    #
    #         self.ui.tW_info.setColumnCount(numcols)
    #         self.ui.tW_info.setRowCount(numrows)
    #
    #         header = self.ui.tW_info.horizontalHeader()
    #         header.setSectionResizeMode(1, QHeaderView.Stretch)
    #         self.ui.tW_info.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
    #
    #         self.ui.tW_info.horizontalHeader().setVisible(False)
    #         self.ui.tW_info.verticalHeader().setVisible(False)
    #
    #         self.ui.tW_info.clicked.connect(self.clicked_fio)
    #
    #         self.ui.tW_info.setItem(0, 0, QTableWidgetItem('Филиал:'))
    #         self.ui.tW_info.setItem(1, 0, QTableWidgetItem('Служба:'))
    #         self.ui.tW_info.setItem(2, 0, QTableWidgetItem('Табельный №:'))
    #         self.ui.tW_info.setItem(3, 0, QTableWidgetItem('ФИО:'))
    #         self.ui.tW_info.setItem(4, 0, QTableWidgetItem('Должность:'))
    #         self.ui.tW_info.setItem(5, 0, QTableWidgetItem('Телефон:'))
    #         self.ui.tW_info.setItem(6, 0, QTableWidgetItem('Почта:'))
    #         self.ui.tW_info.setItem(7, 0, QTableWidgetItem('Расположение:'))
    #         self.ui.tW_info.setItem(8, 0, QTableWidgetItem('Логин:'))
    #         self.ui.tW_info.setItem(9, 0, QTableWidgetItem('ДК:'))
    #         self.ui.tW_info.setItem(10, 0, QTableWidgetItem('АРМ:'))
    #         self.ui.tW_info.setItem(11, 0, QTableWidgetItem('Работает:'))
    #
    #         qt_size = self.getQTableWidgetSize(self.ui.tW_info)
    #         self.ui.tW_info.setMaximumHeight(qt_size.height())
    #         self.ui.tW_info.setMinimumSize(qt_size)
    #
    #         self.checkBox_status = QtWidgets.QCheckBox()
    #         self.ui.tW_info.setCellWidget(11, 1, self.checkBox_status)
    #         self.checkBox_status.setEnabled(False)
    #
    #         for rowPosition in range(numrows):
    #             self.ui.tW_info.item(rowPosition, 0).setBackground(QBrush(Qt.lightGray))  # серый фон
    #             self.ui.tW_info.item(rowPosition, 0).setTextAlignment(
    #                 Qt.AlignVCenter | Qt.AlignRight)  # Выравнивание по правому краю
    #             self.ui.tW_info.item(rowPosition, 0).setFlags(
    #                 Qt.ItemIsEnabled | Qt.ItemIsSelectable)  # запрет редактирования
    #
    #         table = QTableWidgetItem()
    #         table.setTextAlignment(3)
    #
    #         self.ui.tW_info.setEditTriggers(QAbstractItemView.AllEditTriggers)
    #         self.ui.tW_info.setFocusPolicy(Qt.NoFocus)
    #         self.ui.tW_info.setSelectionMode(QAbstractItemView.NoSelection)
    #
    #         collor_tex = QColor('#07437D')
    #         self.ui.tW_info.item(3, 0).setForeground(QBrush(collor_tex))
    #
    #     def create_tW_user():
    #         numrows = 0
    #         numcols = 2
    #
    #         self.ui.tW_list.setColumnCount(numcols)
    #         self.ui.tW_list.setRowCount(numrows)
    #
    #         self.ui.tW_list.setEditTriggers(QAbstractItemView.NoEditTriggers)  # read only
    #         self.ui.tW_list.verticalHeader().setVisible(False)
    #
    #         self.ui.tW_list.setHorizontalHeaderItem(0, QTableWidgetItem('employeeId'))
    #         self.ui.tW_list.setHorizontalHeaderItem(1, QTableWidgetItem('Пользователи'))
    #         self.ui.tW_list.setColumnHidden(0, True)
    #
    #         header = self.ui.tW_list.horizontalHeader()
    #         header.setSectionResizeMode(1, QHeaderView.Stretch)
    #
    #         self.ui.tW_list.setEditTriggers(QAbstractItemView.NoEditTriggers)  # read only
    #         self.ui.tW_list.itemSelectionChanged.connect(self.changed_current_cell_user)
    #
    #     create_item_data()
    #     create_tW_user()
    #
    # def print_tW_user(self, id_user=None):
    #     def load_users():
    #
    #         checked_branch = config['User']['checked_branch']
    #         checked_branch = list(checked_branch)
    #
    #         checked_department = config['User']['checked_department']
    #         checked_department = list(checked_department)
    #
    #         checked_szi = list(map(int, config['User']['checked_szi']))
    #         if Helper_all.convert_bool(config['User']['checkBox_szi']) == True:
    #             checked_szi.append(0)
    #
    #         checked_system = list(map(int, config['User']['checked_system']))
    #         if Helper_all.convert_bool(config['User']['checkBox_system']) == True:
    #             checked_system.append(0)
    #
    #         serch_text = '%' + self.ui.lineEdit_searchUser.text() + '%'
    #         if serch_text == '':
    #             serch_text = '%'
    #
    #         status_work = Helper_all.get_status(Helper_all.convert_bool(config['User']['checkB_statusOn']),
    #                                             Helper_all.convert_bool(config['User']['checkB_statusOff']))
    #
    #         # else:
    #         user = self.s.query(User.employeeId, User.fio).select_from(User). \
    #             join(User_System, isouter=True). \
    #             join(SziFileInst, isouter=True). \
    #             join(SziAccounting, isouter=True). \
    #             filter(or_(
    #             coalesce(SziAccounting.sziType_id, 0).in_((checked_szi))),
    #             coalesce(SziAccounting.status, True) == True). \
    #             filter(coalesce(User_System.id_inf_system, 0).in_((checked_system))). \
    #             filter(User.branch_id.in_((checked_branch))). \
    #             filter(User.departmet_id.in_((checked_department))). \
    #             filter(User.statusWork.like(status_work)). \
    #             filter(or_(
    #             User.employeeId.like(serch_text),
    #             User.fio.like(serch_text),
    #             User.post.like(serch_text),
    #             User.phone.like(serch_text),
    #             User.eMail.like(serch_text),
    #             User.address.like(serch_text),
    #             User.login.like(serch_text),
    #             User.dk.like(serch_text),
    #             User.armName.like(serch_text))). \
    #             group_by(User.employeeId). \
    #             order_by(User.fio)
    #
    #         return user
    #
    #     self.ui.tW_list.setRowCount(0)
    #
    #     users = load_users()
    #
    #     self.ui.statusBar.showMessage('Количество: ' + str(users.count()))
    #
    #     for i in users:
    #         rowPosition = self.ui.tW_list.rowCount()
    #         self.ui.tW_list.insertRow(rowPosition)
    #
    #         self.ui.tW_list.setItem(rowPosition, 0, QTableWidgetItem(str(i[0])))
    #         self.ui.tW_list.setItem(rowPosition, 1, QTableWidgetItem(i[1]))
    #
    #     self.select_row_intable((id_user))
    #
    # def select_row_intable(self, employeeId):
    #     '''Функция выделяющая строку в таблице, ели employeeId=None первую строку,
    #     если есть выделяет указанную'''
    #     if employeeId == None:
    #         if self.ui.tW_list.rowCount() > 0:
    #             self.ui.tW_list.setRangeSelected(QTableWidgetSelectionRange(0, 1, 0, 1), True)
    #             self.print_fill_user(self.ui.tW_list.item(0, 0).text())
    #             self.ui.scrollArea.setVisible(True)
    #         else:
    #             self.ui.scrollArea.setVisible(False)
    #     else:
    #         for i in range(self.ui.tW_list.rowCount()):
    #             if int(self.ui.tW_list.item(i, 0).text()) == int(employeeId):
    #                 self.ui.tW_list.selectRow(i)
    #
    #     # else:
    #
    # def getQTableWidgetSize(self, table):
    #     w = table.verticalHeader().width() - 15  # +4 seems to be needed
    #     for i in range(table.columnCount()):
    #         w += table.columnWidth(i)  # seems to include gridline (on my machine)
    #     h = table.horizontalHeader().height()  # +4 seems to be needed
    #     for i in range(table.rowCount()):
    #         h += table.rowHeight(i)
    #     return QtCore.QSize(w, h)
